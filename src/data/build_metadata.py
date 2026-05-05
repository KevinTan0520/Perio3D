from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from statistics import mean
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook


IQS_SHEET = "IQS"
OE_SHEET = "OE "


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build case-level metadata from SIOP quality workbook.")
    parser.add_argument("--quality-workbook", type=Path, default=Path("dataset/SIOP quality evaluation.xlsx"))
    parser.add_argument("--dataset-root", type=Path, default=Path("dataset"))
    parser.add_argument("--metadata-out", type=Path, default=Path("data/processed/metadata.csv"))
    parser.add_argument("--summary-out", type=Path, default=Path("outputs/reports/metadata_summary.json"))
    return parser.parse_args()


def split_case_uid(case_uid: str) -> tuple[str, str]:
    if "-" not in case_uid:
        return "unknown", case_uid
    return case_uid.split("-", 1)[0], case_uid


def clean_header(value: object) -> str:
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    return (
        text.replace("/", "_")
        .replace(" ", "_")
        .replace("-", "_")
        .replace("(", "")
        .replace(")", "")
        .replace(",", "")
    )


def read_iqs_rows(workbook_path: Path) -> Dict[str, Dict[str, object]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if IQS_SHEET not in wb.sheetnames:
        raise ValueError(f"missing sheet: {IQS_SHEET}")

    ws = wb[IQS_SHEET]
    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    score_names = [clean_header(v) for v in header[1:] if v is not None]

    rows: Dict[str, Dict[str, object]] = {}
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        case_uid = str(raw_row[0]).strip() if raw_row and raw_row[0] is not None else ""
        if not case_uid:
            continue

        scores = []
        row: Dict[str, object] = {"case_uid": case_uid}
        for name, value in zip(score_names, raw_row[1:]):
            if value is None:
                continue
            numeric_value = float(value)
            row[f"iqs_{name}"] = numeric_value
            scores.append(numeric_value)

        if scores:
            row["iqs_num_scores"] = len(scores)
            row["iqs_mean"] = round(mean(scores), 4)
            row["iqs_min"] = min(scores)
            row["iqs_max"] = max(scores)
        else:
            row["iqs_num_scores"] = 0
            row["iqs_mean"] = ""
            row["iqs_min"] = ""
            row["iqs_max"] = ""

        rows[case_uid] = row

    return rows


def read_oe_rows(workbook_path: Path) -> Dict[str, Dict[str, object]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if OE_SHEET not in wb.sheetnames:
        raise ValueError(f"missing sheet: {OE_SHEET!r}")

    ws = wb[OE_SHEET]
    rows: Dict[str, Dict[str, object]] = {}
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        case_uid = str(raw_row[0]).strip() if raw_row and raw_row[0] is not None else ""
        if not case_uid:
            continue
        overall = raw_row[1] if len(raw_row) > 1 else None
        gingival_index = raw_row[2] if len(raw_row) > 2 else None
        rows[case_uid] = {
            "overall_evaluation": "" if overall is None else float(overall),
            "gingival_index": "" if gingival_index is None else float(gingival_index),
        }
    return rows


def count_dataset_images(dataset_root: Path, case_uid: str) -> tuple[int, bool]:
    group_id, case_id = split_case_uid(case_uid)
    case_dir = dataset_root / group_id / case_id
    if not case_dir.exists():
        return 0, False
    image_count = sum(
        1
        for path in case_dir.iterdir()
        if path.is_file()
        and not path.name.startswith("._")
        and path.name != ".DS_Store"
        and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
    )
    return image_count, True


def quality_bucket(iqs_mean: object, overall_evaluation: object) -> str:
    try:
        iqs_value = float(iqs_mean)
        overall_value = float(overall_evaluation)
    except (TypeError, ValueError):
        return "unknown"

    if iqs_value >= 8 and overall_value >= 4:
        return "high"
    if iqs_value >= 6 and overall_value >= 3:
        return "medium"
    return "low"


def write_csv(path: Path, rows: Iterable[Dict[str, object]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def main() -> None:
    args = parse_args()
    if not args.quality_workbook.exists():
        raise FileNotFoundError(f"quality workbook not found: {args.quality_workbook}")

    iqs_rows = read_iqs_rows(args.quality_workbook)
    oe_rows = read_oe_rows(args.quality_workbook)
    case_uids = sorted(set(iqs_rows) | set(oe_rows))

    rows: List[Dict[str, object]] = []
    for case_uid in case_uids:
        group_id, case_id = split_case_uid(case_uid)
        dataset_image_count, dataset_case_exists = count_dataset_images(args.dataset_root, case_uid)
        row: Dict[str, object] = {
            "case_uid": case_uid,
            "group_id": group_id,
            "case_id": case_id,
            "dataset_case_exists": int(dataset_case_exists),
            "dataset_image_count": dataset_image_count,
        }
        row.update(iqs_rows.get(case_uid, {}))
        row.update(oe_rows.get(case_uid, {}))
        row["quality_bucket"] = quality_bucket(row.get("iqs_mean"), row.get("overall_evaluation"))
        rows.append(row)

    fieldnames = [
        "case_uid",
        "group_id",
        "case_id",
        "dataset_case_exists",
        "dataset_image_count",
        "iqs_p_01",
        "iqs_p_02",
        "iqs_p_03",
        "iqs_p_04",
        "iqs_p_05",
        "iqs_p_06",
        "iqs_p_07",
        "iqs_p_08",
        "iqs_p_09",
        "iqs_num_scores",
        "iqs_mean",
        "iqs_min",
        "iqs_max",
        "overall_evaluation",
        "gingival_index",
        "quality_bucket",
    ]
    write_csv(args.metadata_out, rows, fieldnames)

    bucket_counts: Dict[str, int] = {}
    gi_counts: Dict[str, int] = {}
    for row in rows:
        bucket = str(row["quality_bucket"])
        gi = str(row.get("gingival_index", ""))
        bucket_counts[bucket] = bucket_counts.get(bucket, 0) + 1
        gi_counts[gi] = gi_counts.get(gi, 0) + 1

    summary = {
        "num_metadata_cases": len(rows),
        "num_dataset_matched_cases": sum(int(r["dataset_case_exists"]) for r in rows),
        "quality_bucket_counts": bucket_counts,
        "gingival_index_counts": gi_counts,
        "source_workbook": str(args.quality_workbook.as_posix()),
    }
    args.summary_out.parent.mkdir(parents=True, exist_ok=True)
    with args.summary_out.open("w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    print(f"[metadata] metadata written to: {args.metadata_out}")
    print(f"[metadata] summary written to: {args.summary_out}")


if __name__ == "__main__":
    main()
